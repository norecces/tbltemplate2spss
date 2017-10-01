# -*- coding: utf-8 -*-
from __future__ import print_function, division, unicode_literals

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
from collections import OrderedDict
import codecs
import copy

from models import *
from parsers import DUMMY_FIELDS

from parsers import SavFile
import time
import click

VARSTOCASES_SPLIT = '_'


class TemplateMaker(object):

    def __init__(self, template_file_path, survey_structure, treat_as_independent_vars=None):
        self.template_file_path = template_file_path
        self.plain_structure = survey_structure
        self.hierarchical_structure = self.plain_structure.convert_to_hierarchical_structure()
        self.independent_vars = treat_as_independent_vars
        self.varstocases_vars = self._find_varstocases_vars()

    def _find_varstocases_vars(self):
        varstocases_vars = OrderedDict()

        variables_ids = self.hierarchical_structure.get_all_questions_ids()
        variables_ids_split = []
        for variables_id in variables_ids:
            if not self.independent_vars:
                variables_ids_split.append(variables_id)
                continue

            if variables_id not in self.independent_vars:
                variables_ids_split.append(variables_id.rsplit(VARSTOCASES_SPLIT, 1)[0])
            else:
                variables_ids_split.append(variables_id)

        for variable_idx, variable_id_split in enumerate(variables_ids_split):
            if variables_ids[variable_idx].find(VARSTOCASES_SPLIT) == -1:
                continue
            if variable_id_split in varstocases_vars:
                varstocases_vars[variable_id_split].append(variable_idx)
            else:
                varstocases_vars[variable_id_split] = [variable_idx]

        cp_varstocases_vars = copy.deepcopy(varstocases_vars)  # RuntimeError: OrderedDict mutated during iteration
        for variable_id_split in cp_varstocases_vars:
            if len(cp_varstocases_vars[variable_id_split]) < 2:
                varstocases_vars.pop(variable_id_split, None)
        return varstocases_vars

    def download_template(self, path=None):
        """This function parses variable and value labels and output template file to path in xlsx file.
            Manager should fill xlsx file and upload it to server """

        if not path:
            path = self.template_file_path

        wb = Workbook()
        ws = wb.get_active_sheet()

        # filling tables
        ws.title = 'tables'

        for col, lb in enumerate(['QuestionID', 'Variables', 'Title', 'Subtitle\\Question', 'Caption', 'Corner', 'Properties']):
            # making some headers on the list "table"
            cell = ws.cell(row=1, column=col+1)
            cell.value = lb


        # do some nice-looking things with header
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 30

        row = 2
        for question_id in self.hierarchical_structure.get_all_questions_ids():
            # fill sheet by variable labels information
            if question_id in DUMMY_FIELDS:
                continue

            question_structure = self.hierarchical_structure.get_variable_by_id(question_id)
            ws.cell(row=row, column=1).value = question_id
            ws.cell(row=row, column=2).value = ' '.join(question_structure['variable_children'])
            ws.cell(row=row, column=4).value = question_structure['variable_label']
            ws.cell(row=row, column=5).value = u'База: все респонденты'
            row += 1

        wb.create_sheet(title='labels')
        ws = wb.get_sheet_by_name('labels')

        for col, lb in enumerate(['QuestionID', 'Variable', 'Value', 'Label']):
            cell = ws.cell(row=1, column=col+1)
            cell.value = lb
            # cell.fill.fill_type = FILL_SOLID
            # cell.style.fill.start_color.index = 'FFD3D3D3'

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 100
        row = 2

        for question_id in self.hierarchical_structure.get_all_questions_ids():
            if question_id in DUMMY_FIELDS:
                continue
            question_structure = self.hierarchical_structure.get_variable_by_id(question_id)
            ws.cell(row=row, column=1).value = question_id
            ws.cell(row=row, column=2).value = ' '.join(question_structure['variable_children'])
            for (k, v) in question_structure['variable_values'].items():
                ws.cell(row=row, column=3).value = k
                ws.cell(row=row, column=4).value = v
                row += 1
            row += 1

        wb.save(path)

    def upload_template(self, path=None):
        if not path:
            path = self.template_file_path

        wb = load_workbook(filename=path)
        ws = wb.get_sheet_by_name(name='tables')
        row = 2
        tables_set = TablesSet()
        spss_syntax_file_path = os.path.splitext(path)[0] + "_lin.sps"
        spss_syntax_file = codecs.open(spss_syntax_file_path, mode='w', encoding='utf-8', errors='replace')

        spss_lab_file_path = os.path.splitext(path)[0] + "_lab.sps"
        spss_lab_file = codecs.open(spss_lab_file_path, mode='w', encoding='utf-8', errors='replace')

        def _to_unicode(string):
            if string is None:
                return u''
            if isinstance(string, str):
                return string
            return string.encode('uft-8')

        while row <= ws.max_row:

            question_id = ws.cell(row=row, column=1).value
            question_structure = self.hierarchical_structure.get_variable_by_id(question_id)
            question_structure['variable_label'] = ws.cell(row=row, column=4).value

            table = Table(question_structure=question_structure)
            table.id = question_id
            table.title = _to_unicode(ws.cell(row=row, column=3).value)
            table.subtitle = _to_unicode(ws.cell(row=row, column=4).value)
            table.footer = _to_unicode(ws.cell(row=row, column=5).value)
            table.corner = _to_unicode(ws.cell(row=row, column=6).value)
            table.rows = ws.cell(row=row, column=2).value.split(' ')
            table_statistics = TableStatistics()
            table_statistics.add_properties(ws.cell(row=row, column=7).value)
            table.statistics = table_statistics

            tables_set.add_table(table)

            row += 1

        ws = wb.get_sheet_by_name(name='labels')
        row = 2
        previous_question_id = ''
        while row <= ws.max_row:
            question_id = ws.cell(row=row, column=1).value
            if not question_id:
                question_id = previous_question_id
            if question_id:
                question_structure = self.hierarchical_structure.get_variable_by_id(question_id)
                question_structure['variable_values'].update(
                    {ws.cell(row=row, column=3).value: ws.cell(row=row, column=4).value}
                )
            previous_question_id = question_id
            row += 1

        for table in tables_set.tables:
            split_id = table.id.rsplit(VARSTOCASES_SPLIT, 1)[0]
            if split_id in self.varstocases_vars and not table.id.startswith('pre'):
                if len(self.varstocases_vars[split_id]) > 1:
                    spss_syntax_file.write(
                        self._make_varstocases_syntax(
                            variables_idxs=self.varstocases_vars[split_id],
                            tables_set=tables_set
                        )
                    )
                    table.rows = self.hierarchical_structure[self.varstocases_vars[split_id][0]]['variable_children']
                    self.varstocases_vars[split_id] = []
                    spss_syntax_file.write(table
                                           .to_syntax('spss')
                                           .replace(u'tban', u'rot_idx by tban')
                                           .replace(u'sban', u'sban rot_idx')
                    )

                    spss_syntax_file.write(u'\ngetbase.\n')
            else:
                spss_syntax_file.write(table.to_syntax('spss'))

        spss_syntax_file.close()

        for question in self.hierarchical_structure:
            for child in question['variable_children']:
                if question['variable_label'] and question['variable_label'] != '':
                    try:
                        spss_lab_file.write(u'var lab {0} "{1}".\n'.format(
                            child.replace(u'-', u''), question['variable_label'])
                        )
                    except UnicodeDecodeError:
                        print(child)
            if len(question['variable_values']):
                spss_lab_file.write(u'val lab {0}\n'.format(
                    u' '.join([child.replace(u'-', u'') for child in question['variable_children']]))
                )

                try:
                    spss_lab_file.write(u'{0}.\n'.format(
                        u'\n'.join(
                            [u'{k} "{v}"'.format(k=k, v=v) for k, v in question['variable_values'].items() if v]))
                    )
                except UnicodeDecodeError:
                    print(question)
                except TypeError:
                    print(question)

        spss_lab_file.close()

    def _make_varstocases_syntax(self, variables_idxs, tables_set):

        varstocases_text = u''

        variables_list = []
        variables_titles = []
        for variable_idx in variables_idxs:
            variables_list.append(self.hierarchical_structure[variable_idx]['variable_children'])
            variables_titles.append(
                tables_set.get_table_by_id(
                    self.hierarchical_structure[variable_idx]['variable_id']
                ).subtitle
            )

        variables_joined = map(list, zip(*variables_list))

        varstocases_text += u'VARSTOCASES\n'
        for i_idx in range(len(variables_joined)):
                varstocases_text += u'/make ' + variables_joined[i_idx][0] + u' from ' + u' '.join(variables_joined[i_idx]) + u'\n'

        varstocases_text += u'/index rot_idx.\n\n'
        varstocases_text += u'val lab rot_idx\n'
        varstocases_text += u'\n'.join('%s "%s"' % (i+1, v.capitalize()) for i, v in enumerate(variables_titles) if v is not None)
        varstocases_text += u'.\n\n'

        return varstocases_text


def create_template(sav_file_path, multiple_choice_separator='@',
                    use_unlabeled_values=False, template_file_path=None):

    t1 = time.time()
    sav_file = SavFile(sav_file_name=sav_file_path,
                       use_unlabeled_values=use_unlabeled_values,
                       multiple_choice_separator=multiple_choice_separator)

    print("sav file is readed for ", time.time() - t1, 'seconds')

    if not template_file_path:
        template_file_path = sav_file_path + '.xlsx'

    template = TemplateMaker(template_file_path=template_file_path, survey_structure=sav_file.plain_struct)

    print("template initiated")

    return template


def download_xlsx_template(sav_file_path, multiple_choice_separator='@',
                           use_unlabeled_values=False, template_file_path=None):

    xlsx_template = create_template(sav_file_path=sav_file_path,
                                    multiple_choice_separator=multiple_choice_separator,
                                    use_unlabeled_values=use_unlabeled_values,
                                    template_file_path=template_file_path)

    xlsx_template.download_template()
    print('template successfully created at ', xlsx_template.template_file_path)


def upload_xlsx_templae(sav_file_path, multiple_choice_separator='@',
                        use_unlabeled_values=False, template_file_path=None):

    xlsx_template = create_template(sav_file_path=sav_file_path,
                                    multiple_choice_separator=multiple_choice_separator,
                                    use_unlabeled_values=use_unlabeled_values,
                                    template_file_path=template_file_path)

    xlsx_template.upload_template()
    print('spss files successfully created at ', xlsx_template.template_file_path)


@click.command()
@click.argument('action', type=click.Choice(choices=('upload', 'download')))
@click.argument('sav_file_path', type=click.Path(exists=True, readable=True, file_okay=True))
@click.option('--multiple-choice-separator', help='multiple choice separator', default='@', type=str)
@click.option('--xlsx-file-path', help='excel template file path', type=click.Path(writable=True, file_okay=True))
def handle_commands(action, sav_file_path, multiple_choice_separator, xlsx_file_path):

    if action == 'download':
        download_xlsx_template(
            sav_file_path=sav_file_path,
            multiple_choice_separator=multiple_choice_separator,
            template_file_path=xlsx_file_path
        )
    if action == 'upload':
        if xlsx_file_path is None:
            print('please specify xlsx template file path --xlsx-file-path')
            exit()
        upload_xlsx_templae(
            sav_file_path=sav_file_path,
            multiple_choice_separator=multiple_choice_separator,
            template_file_path=xlsx_file_path
        )


if __name__ == '__main__':
    handle_commands()
